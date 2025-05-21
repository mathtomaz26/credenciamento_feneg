
from flask import Flask, request, jsonify, render_template, redirect, session, url_for, make_response
from pymongo import MongoClient
from datetime import datetime
from functools import wraps
import win32print
from threading import Lock
import bcrypt
from datetime import datetime
from flask import Response
import csv
import openpyxl
from io import StringIO
import re


app = Flask(__name__)
app.secret_key = 'chave_super_secreta'

client = MongoClient('mongodb://mongo:27017/')
db = client['credenciamento']
colecao_pre = db['pre_cadastro']
colecao_impressao = db['registro_impressao']
colecao_usuarios = db['usuarios']
indice_impressora = 0
fila_lock = Lock()

def obter_prefixo_impressora():
    config = db['config'].find_one({'chave': 'prefixo_impressora'})
    return config['valor'] if config else 'Brother QL-800'

def atualizar_prefixo_impressora(novo_prefixo):
    db['config'].update_one(
        {'chave': 'prefixo_impressora'},
        {'$set': {'valor': novo_prefixo}},
        upsert=True
    )

def impressoras_disponiveis():
    prefixo = obter_prefixo_impressora()
    return [p[2] for p in win32print.EnumPrinters(2) if prefixo in p[2]]

def proxima_impressora():
    global indice_impressora
    disponiveis = impressoras_disponiveis()
    if not disponiveis:
        raise Exception("Nenhuma impressora disponível com o prefixo configurado.")
    with fila_lock:
        impressora = disponiveis[indice_impressora % len(disponiveis)]
        indice_impressora += 1
    return impressora

def login_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if 'usuario' not in session:
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return wrapper

def admin_required(f):
    @wraps(f)
    def wrapper(*args, **kwargs):
        if session.get('perfil') != 'Administrador':
            return redirect(url_for('credenciamento'))
        return f(*args, **kwargs)
    return wrapper

@app.route('/')
def index():
    if 'usuario' in session:
        return redirect(url_for('credenciamento'))
    return render_template('login.html')

@app.route('/login', methods=['POST'])
def login():
    user = request.form['usuario']
    senha_digitada = request.form['senha']

    usuario = colecao_usuarios.find_one({'user': user})
    if usuario and bcrypt.checkpw(senha_digitada.encode('utf-8'), usuario['password']):
        session['usuario'] = usuario['user']
        session['perfil'] = usuario.get('perfil', 'Comum')

        if session['perfil'] == 'Administrador':
            return redirect(url_for('painel_admin'))
        else:
            return redirect(url_for('credenciamento'))

    return render_template('login.html', erro='Usuário ou senha inválidos')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))

@app.route('/admin')
@login_required
@admin_required
def painel_admin():
    return render_template('admin.html')

@app.route('/credenciamento')
@login_required
def credenciamento():
    return render_template('index.html')

@app.route('/buscar', methods=['POST'])
@login_required
def buscar():
    termo = request.form['termo'].strip().lower()
    resultado = colecao_pre.find_one({
        "$or": [
            {"nome": {"$regex": termo, "$options": "i"}},
            {"cpf":{"$regex": termo, "$options": "i"}},
            {"email": {"$regex": termo, "$options": "i"}}
        ]
    })
    if resultado:
        return jsonify({
            "nome": resultado.get("nome", ""),
            "empresa": resultado.get("empresa", ""),
            "email": resultado.get("email", ""),
            "cpf": resultado.get("cpf", ""),
            "telefone": resultado.get("telefone", "")
        })
    return jsonify({"erro": "Não encontrado"})

@app.route('/imprimir', methods=['POST'])
@login_required
def imprimir():
    nome = request.form['nome']
    empresa = request.form['empresa']
    cpf = re.sub(r'\D', '', request.form.get('cpf', ''))
    email = request.form.get('email', '')
    telefone = request.form.get('telefone', '')
    data_formatada = datetime.now().strftime('%d/%m/%Y')
    data_hora_completa = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    impressora_usada = proxima_impressora()
    imprimir_texto(nome, empresa, data_formatada, impressora_usada)

    colecao_impressao.insert_one({
        "nome": nome,
        "empresa": empresa,
        "CPF": cpf,
        "email": email,
        "telefone": telefone,
        "data_impressao": data_formatada,
        "data_hora_completa": data_hora_completa,
        "impressora": impressora_usada
        
    })

    return jsonify({'status': 'impresso'})

from datetime import datetime  # certifique-se de ter isso no topo do seu servidor

@app.route('/historico')
@login_required
def historico():
    filtro = {}

    nome = request.args.get('nome', '').strip()
    if nome:
        filtro['nome'] = {'$regex': nome, '$options': 'i'}

    impressora = request.args.get('impressora', '').strip()
    if impressora:
        filtro['impressora'] = {'$regex': impressora, '$options': 'i'}

    data = request.args.get('data', '').strip()
    if data:
        try:
            data_formatada = datetime.strptime(data, "%Y-%m-%d").strftime("%d/%m/%Y")
            filtro['data_impressao'] = data_formatada
        except ValueError:
            pass

    registros = [
        (
            r.get('nome', ''),
            r.get('empresa', ''),
            r.get('email', ''),
            r.get('telefone', ''),
            r.get('data_impressao', ''),
            r.get('data_hora_completa', ''),
            r.get('impressora', '')
        )
        for r in colecao_impressao.find(filtro, {'_id': 0})
    ]

    return render_template('historico.html', registros=registros)

    registros = [
        (
            r.get('nome', ''),
            r.get('empresa', ''),
            r.get('email', ''),
            r.get('telefone', ''),
            r.get('data_impressao', ''),
            r.get('data_hora_completa', ''),
            r.get('impressora', '')
        )
        for r in colecao_impressao.find(filtro, {'_id': 0})
    ]

    return render_template('historico.html', registros=registros)
import bcrypt  # certifique-se de importar isso no topo do arquivo

@app.route('/usuarios', methods=['GET', 'POST'])
@login_required
@admin_required
def usuarios():
    mensagem = ""
    if request.method == 'POST':
        user = request.form['user']
        senha_plana = request.form['password']
        perfil = request.form['perfil']

        existente = colecao_usuarios.find_one({"user": user})
        if existente:
            mensagem = "Usuário já existe."
        else:
            senha_hash = bcrypt.hashpw(senha_plana.encode('utf-8'), bcrypt.gensalt())
            colecao_usuarios.insert_one({
                "user": user,
                "password": senha_hash,
                "perfil": perfil
            })
            mensagem = "Usuário cadastrado com sucesso."

    usuarios = list(colecao_usuarios.find({}, {'_id': 0, 'password': 0}))  # oculta a senha da listagem
    return render_template('usuarios.html', usuarios=usuarios, mensagem=mensagem)

@app.route('/impressoras', methods=['GET', 'POST'])
@login_required
@admin_required
def impressoras():
    if request.method == 'POST':
        novo_prefixo = request.form.get('prefixo')
        if novo_prefixo:
            atualizar_prefixo_impressora(novo_prefixo)
        return redirect(url_for('impressoras'))

    prefixo_atual = obter_prefixo_impressora()
    lista = impressoras_disponiveis()
    return render_template('impressoras.html', impressoras=lista, prefixo=prefixo_atual)


def imprimir_texto(nome, empresa, data_str, printer_name):
    epl = f'N\nA100,60,0,4,1,1,N,"{nome}"\nA100,120,0,3,1,1,N,"{empresa}"\nA100,180,0,2,1,1,N,"{data_str}"\nP1\n'
    hprinter = win32print.OpenPrinter(printer_name)
    job = win32print.StartDocPrinter(hprinter, 1, ("Etiqueta Elgin", None, "RAW"))
    win32print.StartPagePrinter(hprinter)
    win32print.WritePrinter(hprinter, epl.encode("utf-8"))
    win32print.EndPagePrinter(hprinter)
    win32print.EndDocPrinter(hprinter)
    win32print.ClosePrinter(hprinter)
    
@app.route('/usuarios/editar/<usuario>', methods=['GET', 'POST'])
@login_required
@admin_required
def editar_usuario(usuario):
    if request.method == 'POST':
        nova_senha = request.form['password']
        novo_perfil = request.form['perfil']
        colecao_usuarios.update_one(
            {'user': usuario},
            {'$set': {'password': nova_senha, 'perfil': novo_perfil}}
        )
        return redirect(url_for('usuarios'))

    dados = colecao_usuarios.find_one({'user': usuario}, {'_id': 0})
    if not dados:
        return redirect(url_for('usuarios'))

    return render_template('editar_usuario.html', user=usuario, perfil=dados['perfil'])

@app.route('/usuarios/excluir/<usuario>')
@login_required
@admin_required
def excluir_usuario(usuario):
    colecao_usuarios.delete_one({'user': usuario})
    return redirect(url_for('usuarios'))

@app.route('/exportar-historico')
@login_required
@admin_required
def exportar_historico():
    registros = list(colecao_impressao.find({}, {'_id': 0}))
    output = StringIO()
    writer = csv.DictWriter(output, fieldnames=["nome", "empresa", "email", "telefone", "data_impressao", "data_hora_completa", "impressora"])
    writer.writeheader()
    for r in registros:
        writer.writerow({
            "nome": r.get("nome", ""),
            "empresa": r.get("empresa", ""),
            "email": r.get("email", ""),
            "telefone": r.get("telefone", ""),
            "data_impressao": r.get("data_impressao", ""),
            "data_hora_completa": r.get("data_hora_completa", ""),
            "impressora": r.get("impressora", "")
        })

    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=historico_impressao.csv"
    return response

@app.route('/historico/atualizar')
@login_required
def historico_atualizar():
    import re
    filtro = {}

    nome = request.args.get('nome', '').strip()
    if nome:
        filtro['nome'] = {'$regex': nome, '$options': 'i'}

    impressora = request.args.get('impressora', '').strip()
    if impressora:
        filtro['impressora'] = {'$regex': impressora, '$options': 'i'}

    cpf = request.args.get('cpf', '').strip()
    if cpf:
        cpf = re.sub(r'\D', '', cpf)  # remove tudo que não for dígito
        filtro['CPF'] = {'$regex': cpf, '$options': 'i'}

    data = request.args.get('data', '').strip()
    if data:
        try:
            data_formatada = datetime.strptime(data, "%Y-%m-%d").strftime("%d/%m/%Y")
            filtro['data_impressao'] = data_formatada
        except ValueError:
            pass

    registros = list(colecao_impressao.find(filtro, {'_id': 0}).sort('data_hora_completa', -1))
    return jsonify(registros)

def painel_impressoras():
    return render_template('painel_impressoras.html')

@app.route('/painel-impressoras')
@login_required
@admin_required
def painel_impressoras():
    return render_template('painel_impressoras.html')

@app.route('/painel-impressoras/dados')
@login_required
@admin_required
def painel_dados():
    pipeline = [
        {"$group": {"_id": "$impressora", "total": {"$sum": 1}}},
        {"$project": {"impressora": "$_id", "total": 1, "_id": 0}},
        {"$sort": {"total": -1}}
    ]
    dados = list(colecao_impressao.aggregate(pipeline))
    return jsonify(dados)

@app.route('/upload-pre-cadastro', methods=['GET', 'POST'])
@login_required
@admin_required
def upload_pre_cadastro():
    if request.method == 'POST':
        file = request.files['arquivo']
        if file.filename.endswith('.xlsx'):
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [cell.value for cell in ws[1]]
            dados = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                item = dict(zip(headers, row))
                # Normaliza todos os valores: se for None, vira string vazia
                item = {k: (str(v).strip() if v is not None else "") for k, v in item.items()}
                item["cpf"] = re.sub(r'\D', '', item.get("cpf", ""))
                dados.append(item)
            colecao_pre.insert_many(dados)
            return "Base importada com sucesso!"
        return "Arquivo inválido."
    return render_template('upload_pre_cadastro.html')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
